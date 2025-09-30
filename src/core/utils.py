from pathlib import Path
from typing import Union, Any

from openpyxl import Workbook


def _norm_header(s: Any) -> str:
    """Нормализация заголовка: str, lower, strip. Пустые -> ''."""
    if s is None:
        return ""
    return str(s).strip().lower()


def get_sheet_name(wb, names: Union[list[str], tuple[str, ...]]):
    """Возвращает лист по имени (без учёта регистра).
    Поддерживает несколько возможных названий.
    """
    mapping = {s.lower(): s for s in wb.sheetnames}

    if isinstance(names, str):
        key = names.lower()
        if key in mapping:
            return wb[mapping[key]]
        raise KeyError(f"В книге нет листа '{names}'. Есть: {wb.sheetnames}")

    for nm in names:
        if nm is None:
            continue
        key = str(nm).lower()
        if key in mapping:
            return wb[mapping[key]]

    raise KeyError(f"В книге нет ни одного из листов {list(names)}. Есть: {wb.sheetnames}")


def make_output_path(src_wb_path: Path) -> Path:
    """Возвращает путь для итогового файла рядом с исходником:
    '<dir>/(заполнено) <оригинальное имя>.xlsx'
    """
    prefix = "(заполнено) "
    return src_wb_path.with_name(f"{prefix}{src_wb_path.name}")


def ensure_ws(wb: Workbook, sheet_name: str):
    """Возвращает рабочий лист по имени (без учёта регистра).
    Если листа нет — создаёт новый и добавляет его в книгу.
    """
    existing_map = {s.lower(): s for s in wb.sheetnames}
    key = sheet_name.lower()
    if key not in existing_map:
        ws = wb.create_sheet(sheet_name)
        existing_map[key] = sheet_name
        return ws
    return wb[existing_map[key]]
