from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable, Optional, Union

from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet

from excel_manager.core.row_filters import filter_rows
from excel_manager.core.row_reader import read_rows
from excel_manager.core.utils import get_sheet_name, ensure_ws, _norm_header


StrOrInt = Union[str, int]


@dataclass
class HeaderInfo:
    row_idx: int                 # номер строки заголовка (1-based, как в Excel)
    names: list[str]             # исходные заголовки (как есть)
    name_to_idx: dict[str, int]  # нормализованное имя -> 0-based индекс столбца


class ExcelManager:
    """
    Базовый класс: инициализация книг/листов, заголовки, доступ к значениям,
    фильтрация, копирование столбцов, запись в другую книгу/лист.
    """

    def __init__(
        self,
        path: Union[str, Path],
        sheet: Union[str, int, Iterable[str]] = 0,
        header_row: Optional[int] = None,
        read_only: bool = True,
        data_only: bool = True,
    ):
        """
        path       : путь к XLSX
        sheet      : имя листа (str), индекс (int) или список возможных имён (Iterable[str])
        header_row : если известен номер строки заголовка; иначе будет определён автоматически
        read_only  : открыть в режиме чтения
        data_only  : подставлять вычисленные значения формул
        """
        self.path = Path(path)
        self.wb: Workbook = load_workbook(self.path, read_only=read_only, data_only=data_only)

        # выбор листа
        if isinstance(sheet, (str, list, tuple)):
            self.ws: Worksheet = get_sheet_name(self.wb, sheet)
        elif isinstance(sheet, int):
            try:
                self.ws: Worksheet = self.wb.worksheets[sheet]
            except IndexError:
                raise ValueError(f"В книге нет листа с индексом {sheet}")
        else:
            self.ws: Worksheet = self.wb.worksheets[0]

        # определение заголовка
        self.header: HeaderInfo = self.build_header()

    # ---------- базовые служебные вещи ----------

    def _detect_header_row(self, scan_rows: int = 20) -> int:
        """Ищем первую строку среди первых N, где есть >=2 непустых ячейки.
        Возвращаем её индекс (Excel 1-based). Если не нашли — поднимаем ошибку.
        """
        for i, row in enumerate(
            self.ws.iter_rows(min_row=1, max_row=scan_rows, values_only=True), start=1
        ):
            non_empty = [c for c in row if c not in (None, "", " ")]
            if len(non_empty) >= 2:
                return i

        raise ValueError(
            f"Не удалось определить заголовок на листе '{self.ws.title}'. "
            f"Проверьте первые {scan_rows} строк."
        )

    def build_header(self, header_row: Optional[int] = None) -> HeaderInfo:
        row_idx = header_row or self._detect_header_row()

        rows = list(
            self.ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True)
        )
        if not rows:
            raise ValueError(
                f"В листе '{self.ws.title}' нет строки с индексом {row_idx}. "
                f"Доступно строк: {self.ws.max_row}"
            )

        header_cells = rows[0]
        names = [c if c is not None else "" for c in header_cells]
        name_to_idx = {_norm_header(v): i for i, v in enumerate(names)}

        return HeaderInfo(row_idx=row_idx, names=names, name_to_idx=name_to_idx)

    def list_sheets(self) -> list[str]:
        """Имена листов книги."""
        return list(self.wb.sheetnames)

    def data_rows(self) -> list[list[Any]]:
        """
        Строки данных после заголовка (пропуская полностью пустые).
        Реюз твоей функции read_rows с min_row = header_row + 1.  :contentReference[oaicite:10]{index=10}
        """
        if self._rows_cache is None:
            start = max(self.header.row_idx + 1, 1)
            self._rows_cache = read_rows(self.ws, start_row=start)
        return self._rows_cache

    def count_rows(self) -> int:
        """Количество непустых строк данных (после заголовка)."""
        return len(self.data_rows())

    def headers(self, as_indexed: bool = False) -> list[Union[str, tuple[int, str]]]:
        """
        Возвращает заголовки как список. Если as_indexed=True — [(0, 'A'), (1, 'B'), ...].
        """
        if as_indexed:
            return [(i, v if v is not None else "") for i, v in enumerate(self.header.names)]
        return [v if v is not None else "" for v in self.header.names]

    # ---------- доступ к значениям ----------

    def _col_to_idx(self, col: StrOrInt) -> int:
        """Преобразование 'ИмяКолонки' -> 0-based idx, либо int -> int."""
        if isinstance(col, int):
            return col
        key = _norm_header(col)
        if key not in self.header.name_to_idx:
            raise KeyError(f"Колонка '{col}' не найдена среди заголовков: {self.header.names}")
        return self.header.name_to_idx[key]

    def get_value(self, row_no: int, col: StrOrInt, absolute: bool = False) -> Any:
        """
        Получить значение:
          row_no   : номер строки (1-based). Если absolute=False — это «номер строки данных»
                     (т.е. 1 соответствует первой строке после заголовка).
                     Если absolute=True — это реальный Excel-ряд.
          col      : индекс (0-based) или название столбца.
        """
        col_idx = self._col_to_idx(col)

        if absolute:
            values = next(self.ws.iter_rows(min_row=row_no, max_row=row_no, values_only=True))
            return values[col_idx] if col_idx < len(values) else None

        # относительный к данным
        rows = self.data_rows()
        if not (1 <= row_no <= len(rows)):
            return None
        row = rows[row_no - 1]
        return row[col_idx] if col_idx < len(row) else None

    def get_column_values(self, col: StrOrInt, include_header: bool = False) -> list[Any]:
        """Получить все значения столбца (по индексу или имени)."""
        idx = self._col_to_idx(col)
        if include_header:
            return [self.header.names[idx]] + [r[idx] if idx < len(r) else None for r in self.data_rows()]
        return [r[idx] if idx < len(r) else None for r in self.data_rows()]

    # ---------- фильтрация ----------

    def filter(self, rules: dict[StrOrInt, dict[str, Any]]) -> list[list[Any]]:
        """
        Фильтрация по правилам (как в твоём `filter_rows`), но можно указывать
        колонки по имени или индексу. Возвращает НОВЫЙ список строк.
        Пример rules:
            {
              "статус": {"equals": ["Отменено", "Черновик"]},
              5:       {"contains": ["VIP"]},
            }
        """
        # переводим ключи правил в 0-based индексы
        idx_rules: dict[int, dict[str, Any]] = {}
        for col, cond in rules.items():
            idx_rules[self._col_to_idx(col)] = cond

        # работаем только по данным (после заголовка)
        rows = self.data_rows()
        return filter_rows(rows, idx_rules)

    # ---------- копирование/запись ----------

    def copy_columns(
        self,
        dest_path: Union[str, Path],
        dest_sheet: str,
        columns: list[StrOrInt],
        include_header: bool = True,
        start_cell: str = "A1",
    ) -> Path:
        """
        Скопировать выбранные столбцы текущего листа в ДРУГУЮ книгу/лист.
        Если файла нет — создаём; если листа нет — создаём (регистронезависимо).
        """
        dest_path = Path(dest_path)
        if dest_path.exists():
            dwb = load_workbook(dest_path)
        else:
            dwb = Workbook()
            # openpyxl создаёт один лист "Sheet" по умолчанию; оставим/переименуем при записи

        dws = ensure_ws(dwb, dest_sheet)  # создаст если нет (без учёта регистра)  :contentReference[oaicite:12]{index=12}

        # Готовим таблицу: возможно, с заголовком
        col_indices = [self._col_to_idx(c) for c in columns]
        out_rows: list[list[Any]] = []

        if include_header:
            out_rows.append([self.header.names[i] if i < len(self.header.names) else None for i in col_indices])

        for r in self.data_rows():
            out_rows.append([r[i] if i < len(r) else None for i in col_indices])

        # Куда писать
        from openpyxl.utils import coordinate_to_tuple
        start_row, start_col = coordinate_to_tuple(start_cell)

        for i, row in enumerate(out_rows, start=start_row):
            for j, val in enumerate(row, start=start_col):
                dws.cell(row=i, column=j, value=val)

        dwb.save(dest_path)
        return dest_path

    def write_rows(
        self,
        dest_path: Union[str, Path],
        dest_sheet: str,
        rows: list[list[Any]],
        start_cell: str = "A1",
    ) -> Path:
        """
        Универсальная запись произвольных rows в целевую книгу/лист.
        """
        dest_path = Path(dest_path)
        if dest_path.exists():
            dwb = load_workbook(dest_path)
        else:
            dwb = Workbook()

        dws = ensure_ws(dwb, dest_sheet)  # создаст при отсутствии  :contentReference[oaicite:13]{index=13}

        from openpyxl.utils import coordinate_to_tuple
        start_row, start_col = coordinate_to_tuple(start_cell)

        for i, row in enumerate(rows, start=start_row):
            for j, val in enumerate(row, start=start_col):
                dws.cell(row=i, column=j, value=val)

        dwb.save(dest_path)
        return dest_path

    def transfer_by_headers(
        self,
        dest_path: Union[str, Path],
        dest_sheet: str,
        headers: list[str],
        include_header: bool = True,
        start_cell: str = "A1",
    ) -> Path:
        """
        «Передача названий столбцов и перенос их в другую таблицу»:
        принимаем список заголовков (в нужном порядке), копируем соответствующие колонки.
        """
        return self.copy_columns(
            dest_path=dest_path,
            dest_sheet=dest_sheet,
            columns=headers,
            include_header=include_header,
            start_cell=start_cell,
        )


class ExcelInspector(ExcelManager):
    """
    «Служебный» класс (от основного):
    - вывести заголовки
    - вывести значения из столбца
    - посчитать строки
    - показать листы
    """

    def print_headers(self) -> None:
        print("Заголовки:")
        for i, name in enumerate(self.headers(), start=1):
            print(f"[{i:>2}] {name}")

    def print_sheets(self) -> None:
        print("Листы книги:")
        for i, name in enumerate(self.list_sheets(), start=1):
            print(f"[{i:>2}] {name}")

    def print_column(self, col: StrOrInt, limit: Optional[int] = 20, include_header: bool = False) -> None:
        vals = self.get_column_values(col, include_header=include_header)
        if limit is not None:
            vals = vals[:limit]
        title = f"Столбец: {col!r}"
        print(title)
        print("-" * len(title))
        for i, v in enumerate(vals, start=1):
            print(f"{i:>5}: {v!r}")
